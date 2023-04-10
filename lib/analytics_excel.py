import openpyxl

class open_excel():
    def __init__(self):
        self.excel_path = None
        #self.sheet_name = None

    def read_sheet(self):
        return openpyxl.load_workbook(self.excel_path, read_only=False).sheetnames

    def read_excel(self, sheet_name):
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=False)
            ws = wb.worksheets[sheet_name]
            return ws
        except Exception as e:
            print(e)
            return None
    
    #def get_rowname(self,)